import pandas as pd
import numpy
import os
import openpyxl

states = ["West Virginia"]

important_variables = [
    'custody_49under', 'custody_50to64', 'custody_65over', 'custody_18to24', 'custody_25to44',
    'custody_45to49', 'custody_50plus', 'custody_age_unknown', 'custody_male', 'custody_female',
    'custody_gender_unknown', 'custody_race1_white', 'custody_race1_black', 'custody_race1_latino',
    'custody_race1_other', 'custody_race1_unknown', 'custody_race2_white', 'custody_race2_black',
    'custody_race2_other', 'custody_race2_unknown', 'mortality_49under', 'mortality_50to64',
    'mortality_65over', 'mortality_18to24', 'mortality_25to44', 'mortality_45to49', 'mortality_50plus',
    'mortality_age_unknown', 'mortality_male', 'mortality_female', 'mortality_gender_unknown',
    'mortality_race1_white', 'mortality_race1_black', 'mortality_race1_latino', 'mortality_race1_other',
    'mortality_race1_unknown', 'mortality_race2_white', 'mortality_race2_black', 'mortality_race2_other',
    'mortality_race2_unknown'
]

wb = openpyxl.load_workbook('QC standard File.xlsx')
pandas_type = pd.read_excel('QC standard File.xlsx', sheet_name='State-Year')
ws = wb['State-Year']

def check_authenticity(state, spreadsheet):
    account_for_both = pd.ExcelFile(f'{state}_Univariate.xlsx')
    if spreadsheet not in account_for_both.sheet_names:
        goodbye = open(f'{state}-State-Year.txt', 'w')
        goodbye.write(f'{spreadsheet} DOES NOT EXIST')
        return True
    else:
        return False


def state_year_general(state):
    entire_general_notes = ""
    entire_custody_total = ""
    entire_custody_49under = ""
    entire_custody_50to64 = ""
    entire_custody_65over = ""
    entire_custody_18to24 = ""
    entire_custody_25to44 = ""
    entire_custody_45to49 = ""
    entire_custody_50plus = ""
    entire_custody_age_unknown = ""
    entire_custody_male = ""
    entire_custody_female = ""
    entire_custody_gender_unknown = ""
    entire_custody_race1_white = ""
    entire_custody_race1_black = ""
    entire_custody_race1_latino = ""
    entire_custody_race1_other = ""
    entire_custody_race1_unknown = ""
    entire_custody_race2_white = ""
    entire_custody_race2_black = ""
    entire_custody_race2_other = ""
    entire_custody_race2_unknown = ""
    entire_mortality_49under = ""
    entire_mortality_50to64 = ""
    entire_mortality_65over = ""
    entire_mortality_18to24 = ""
    entire_mortality_25to44 = ""
    entire_mortality_45to49 = ""
    entire_mortality_50plus = ""
    entire_mortality_age_unknown = ""
    entire_mortality_male = ""
    entire_mortality_female = ""
    entire_mortality_gender_unknown = ""
    entire_mortality_race1_white = ""
    entire_mortality_race1_black = ""
    entire_mortality_race1_latino = ""
    entire_mortality_race1_other = ""
    entire_mortality_race1_unknown = ""
    entire_mortality_race2_white = ""
    entire_mortality_race2_black = ""
    entire_mortality_race2_other = ""
    entire_mortality_race2_unknown = ""
    row_index = pandas_type[pandas_type["State"] == state].index[0] + 2
    univariate_state_year = pd.read_excel(f"{state}_Univariate.xlsx", 'State-Year')
    raw_state_year = pd.read_excel(f"{state}.xlsx", 'State-Year')
    univariate_state_year_pointer = univariate_state_year.fillna('no value')
    raw_state_year_pointer = raw_state_year.fillna('no value')
    years_in_file = (univariate_state_year_pointer['year'].tolist())
    year_index = {}
    year_issues = []
    year_unincluded = []
    goodbye = open(f'{state}-State-Year.txt', 'w')
    for years in years_in_file:
        if 2013 <= years <= 2022:
            a_list = univariate_state_year_pointer.index[univariate_state_year_pointer['year'] == years].tolist()
            for index in a_list:
                year_index.update({years: index})
        elif years < 2013 or years > 2022:
            year_issues.append(years)
    for year in range(2013, 2023):
        if year not in years_in_file:
            year_unincluded.append(year)
    for indexed_year in year_index:
        retrieve_again = year_index.get(indexed_year)
        general_message = univariate_state_year_pointer.loc[retrieve_again, 'both_diff']
        message_placeholder = f'{indexed_year}'
        message_placeholder += f" {general_message},"
        for univariate_year in univariate_state_year_pointer.columns:
            the_value = univariate_state_year_pointer.loc[retrieve_again, univariate_year]
            if the_value != 0 and isinstance(the_value,
                                             numpy.int64) and univariate_year != 'c_age_check_diff' and univariate_year != 'c_all_diff' and univariate_year != 'm_age_check_diff' and univariate_year != 'm_all_diff' and univariate_year != 'year':
                message_placeholder += f' {univariate_year} != 0,'
        for raw_year in raw_state_year_pointer.columns:
            raw_checker = raw_state_year_pointer.loc[retrieve_again, raw_year]
            if raw_checker == 'no value':
                if raw_year == 'custody_tot':
                    message_placeholder += f' NO CUSTODY TOTAL PRESENT.'
                    entire_custody_total += f'{indexed_year}\n\n'
                elif raw_year in important_variables:
                    if raw_year == 'custody_49under':
                        entire_custody_49under += f'{indexed_year}\n\n'

                    if raw_year == 'custody_50to64':
                        entire_custody_50to64 += f'{indexed_year}\n\n'

                    if raw_year == 'custody_65over':
                        entire_custody_65over += f'{indexed_year}\n\n'

                    if raw_year == 'custody_18to24':
                        entire_custody_18to24 += f'{indexed_year}\n\n'

                    if raw_year == 'custody_25to44':
                        entire_custody_25to44 += f'{indexed_year}\n\n'

                    if raw_year == 'custody_45to49':
                        entire_custody_45to49 += f'{indexed_year}\n\n'

                    if raw_year == 'custody_50plus':
                        entire_custody_50plus += f'{indexed_year}\n\n'

                    if raw_year == 'custody_age_unknown':
                        entire_custody_age_unknown += f'{indexed_year}\n\n'

                    if raw_year == 'custody_male':
                        entire_custody_male += f'{indexed_year}\n\n'

                    if raw_year == 'custody_female':
                        entire_custody_female += f'{indexed_year}\n\n'

                    if raw_year == 'custody_gender_unknown':
                        entire_custody_gender_unknown += f'{indexed_year}\n\n'

                    if raw_year == 'custody_race1_white':
                        entire_custody_race1_white += f'{indexed_year}\n\n'

                    if raw_year == 'custody_race1_black':
                        entire_custody_race1_black += f'{indexed_year}\n\n'

                    if raw_year == 'custody_race1_latino':
                        entire_custody_race1_latino += f'{indexed_year}\n\n'

                    if raw_year == 'custody_race1_other':
                        entire_custody_race1_other += f'{indexed_year}\n\n'

                    if raw_year == 'custody_race1_unknown':
                        entire_custody_race1_unknown += f'{indexed_year}\n\n'

                    if raw_year == 'custody_race2_white':
                        entire_custody_race2_white += f'{indexed_year}\n\n'

                    if raw_year == 'custody_race2_black':
                        entire_custody_race2_black += f'{indexed_year}\n\n'

                    if raw_year == 'custody_race2_other':
                        entire_custody_race2_other += f'{indexed_year}\n\n'

                    if raw_year == 'custody_race2_unknown':
                        entire_custody_race2_unknown += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_49under':
                        entire_mortality_49under += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_50to64':
                        entire_mortality_50to64 += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_65over':
                        entire_mortality_65over += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_18to24':
                        entire_mortality_18to24 += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_25to44':
                        entire_mortality_25to44 += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_45to49':
                        entire_mortality_45to49 += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_50plus':
                        entire_mortality_50plus += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_age_unknown':
                        entire_mortality_age_unknown += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_male':
                        entire_mortality_male += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_female':
                        entire_mortality_female += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_gender_unknown':
                        entire_mortality_gender_unknown += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_race1_white':
                        entire_mortality_race1_white += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_race1_black':
                        entire_mortality_race1_black += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_race1_latino':
                        entire_mortality_race1_latino += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_race1_other':
                        entire_mortality_race1_other += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_race1_unknown':
                        entire_mortality_race1_unknown += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_race2_white':
                        entire_mortality_race2_white += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_race2_black':
                        entire_mortality_race2_black += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_race2_other':
                        entire_mortality_race2_other += f'{indexed_year}\n\n'

                    if raw_year == 'mortality_race2_unknown':
                        entire_mortality_race2_unknown += f'{indexed_year}\n\n'
                    message_placeholder += f' missing {raw_year},'
                    # globals()[year_variable] += f"{indexed_year}\n\n"  # Updates the variable dynamically
        goodbye.write(f'{message_placeholder}\n\n')
        entire_general_notes += f'{message_placeholder}\n\n'
    for second_time in year_issues:
        goodbye.write(f'{second_time} is included\n\n')
    for third_time in year_unincluded:
        goodbye.write(f'{third_time} is missing\n\n')
    write_into = ws.cell(row_index, 3)
    write_into.value = f'{entire_general_notes}'
    no_custody_cell = ws.cell(row=row_index, column=pandas_type.columns.get_loc('missing custody_total') + 1)
    no_custody_cell.value = entire_custody_total

    write_into_custody_49under = ws.cell(row=row_index,
                                         column=pandas_type.columns.get_loc('missing custody_49under') + 1)
    write_into_custody_49under.value = entire_custody_49under

    write_into_custody_50to64 = ws.cell(row=row_index, column=pandas_type.columns.get_loc('missing custody_50to64') + 1)
    write_into_custody_50to64.value = entire_custody_50to64

    write_into_custody_65over = ws.cell(row=row_index, column=pandas_type.columns.get_loc('missing custody_65over') + 1)
    write_into_custody_65over.value = entire_custody_65over

    write_into_custody_18to24 = ws.cell(row=row_index, column=pandas_type.columns.get_loc('missing custody_18to24') + 1)
    write_into_custody_18to24.value = entire_custody_18to24

    write_into_custody_25to44 = ws.cell(row=row_index, column=pandas_type.columns.get_loc('missing custody_25to44') + 1)
    write_into_custody_25to44.value = entire_custody_25to44

    write_into_custody_45to49 = ws.cell(row=row_index, column=pandas_type.columns.get_loc('missing custody_45to49') + 1)
    write_into_custody_45to49.value = entire_custody_45to49

    write_into_custody_50plus = ws.cell(row=row_index, column=pandas_type.columns.get_loc('missing custody_50plus') + 1)
    write_into_custody_50plus.value = entire_custody_50plus

    write_into_custody_age_unknown = ws.cell(row=row_index,
                                             column=pandas_type.columns.get_loc('missing custody_age_unknown') + 1)
    write_into_custody_age_unknown.value = entire_custody_age_unknown

    write_into_custody_male = ws.cell(row=row_index, column=pandas_type.columns.get_loc('missing custody_male') + 1)
    write_into_custody_male.value = entire_custody_male

    write_into_custody_female = ws.cell(row=row_index, column=pandas_type.columns.get_loc('missing custody_female') + 1)
    write_into_custody_female.value = entire_custody_female

    write_into_custody_gender_unknown = ws.cell(row=row_index, column=pandas_type.columns.get_loc(
        'missing custody_gender_unknown') + 1)
    write_into_custody_gender_unknown.value = entire_custody_gender_unknown

    write_into_custody_race1_white = ws.cell(row=row_index,
                                             column=pandas_type.columns.get_loc('missing custody_race1_white') + 1)
    write_into_custody_race1_white.value = entire_custody_race1_white

    write_into_custody_race1_black = ws.cell(row=row_index,
                                             column=pandas_type.columns.get_loc('missing custody_race1_black') + 1)
    write_into_custody_race1_black.value = entire_custody_race1_black

    write_into_custody_race1_latino = ws.cell(row=row_index,
                                              column=pandas_type.columns.get_loc('missing custody_race1_latino') + 1)
    write_into_custody_race1_latino.value = entire_custody_race1_latino

    write_into_custody_race1_other = ws.cell(row=row_index,
                                             column=pandas_type.columns.get_loc('missing custody_race1_other') + 1)
    write_into_custody_race1_other.value = entire_custody_race1_other

    write_into_custody_race1_unknown = ws.cell(row=row_index,
                                               column=pandas_type.columns.get_loc('missing custody_race1_unknown') + 1)
    write_into_custody_race1_unknown.value = entire_custody_race1_unknown

    write_into_custody_race2_white = ws.cell(row=row_index,
                                             column=pandas_type.columns.get_loc('missing custody_race2_white') + 1)
    write_into_custody_race2_white.value = entire_custody_race2_white

    write_into_custody_race2_black = ws.cell(row=row_index,
                                             column=pandas_type.columns.get_loc('missing custody_race2_black') + 1)
    write_into_custody_race2_black.value = entire_custody_race2_black

    write_into_custody_race2_other = ws.cell(row=row_index,
                                             column=pandas_type.columns.get_loc('missing custody_race2_other') + 1)
    write_into_custody_race2_other.value = entire_custody_race2_other

    write_into_custody_race2_unknown = ws.cell(row=row_index,
                                               column=pandas_type.columns.get_loc('missing custody_race2_unknown') + 1)
    write_into_custody_race2_unknown.value = entire_custody_race2_unknown

    write_into_mortality_49under = ws.cell(row=row_index,
                                           column=pandas_type.columns.get_loc('missing mortality_49under') + 1)
    write_into_mortality_49under.value = entire_mortality_49under

    write_into_mortality_50to64 = ws.cell(row=row_index,
                                          column=pandas_type.columns.get_loc('missing mortality_50to64') + 1)
    write_into_mortality_50to64.value = entire_mortality_50to64

    write_into_mortality_65over = ws.cell(row=row_index,
                                          column=pandas_type.columns.get_loc('missing mortality_65over') + 1)
    write_into_mortality_65over.value = entire_mortality_65over

    write_into_mortality_18to24 = ws.cell(row=row_index,
                                          column=pandas_type.columns.get_loc('missing mortality_18to24') + 1)
    write_into_mortality_18to24.value = entire_mortality_18to24

    write_into_mortality_25to44 = ws.cell(row=row_index,
                                          column=pandas_type.columns.get_loc('missing mortality_25to44') + 1)
    write_into_mortality_25to44.value = entire_mortality_25to44

    write_into_mortality_45to49 = ws.cell(row=row_index,
                                          column=pandas_type.columns.get_loc('missing mortality_45to49') + 1)
    write_into_mortality_45to49.value = entire_mortality_45to49

    write_into_mortality_50plus = ws.cell(row=row_index,
                                          column=pandas_type.columns.get_loc('missing mortality_50plus') + 1)
    write_into_mortality_50plus.value = entire_mortality_50plus

    write_into_mortality_age_unknown = ws.cell(row=row_index,
                                               column=pandas_type.columns.get_loc('missing mortality_age_unknown') + 1)
    write_into_mortality_age_unknown.value = entire_mortality_age_unknown

    write_into_mortality_male = ws.cell(row=row_index, column=pandas_type.columns.get_loc('missing mortality_male') + 1)
    write_into_mortality_male.value = entire_mortality_male

    write_into_mortality_female = ws.cell(row=row_index,
                                          column=pandas_type.columns.get_loc('missing mortality_female') + 1)
    write_into_mortality_female.value = entire_mortality_female

    write_into_mortality_gender_unknown = ws.cell(row=row_index, column=pandas_type.columns.get_loc(
        'missing mortality_gender_unknown') + 1)
    write_into_mortality_gender_unknown.value = entire_mortality_gender_unknown

    write_into_mortality_race1_white = ws.cell(row=row_index,
                                               column=pandas_type.columns.get_loc('missing mortality_race1_white') + 1)
    write_into_mortality_race1_white.value = entire_mortality_race1_white

    write_into_mortality_race1_black = ws.cell(row=row_index,
                                               column=pandas_type.columns.get_loc('missing mortality_race1_black') + 1)
    write_into_mortality_race1_black.value = entire_mortality_race1_black

    write_into_mortality_race1_latino = ws.cell(row=row_index, column=pandas_type.columns.get_loc(
        'missing mortality_race1_latino') + 1)
    write_into_mortality_race1_latino.value = entire_mortality_race1_latino

    write_into_mortality_race1_other = ws.cell(row=row_index,
                                               column=pandas_type.columns.get_loc('missing mortality_race1_other') + 1)
    write_into_mortality_race1_other.value = entire_mortality_race1_other

    write_into_mortality_race1_unknown = ws.cell(row=row_index, column=pandas_type.columns.get_loc(
        'missing mortality_race1_unknown') + 1)
    write_into_mortality_race1_unknown.value = entire_mortality_race1_unknown

    write_into_mortality_race2_white = ws.cell(row=row_index,
                                               column=pandas_type.columns.get_loc('missing mortality_race2_white') + 1)
    write_into_mortality_race2_white.value = entire_mortality_race2_white

    write_into_mortality_race2_black = ws.cell(row=row_index,
                                               column=pandas_type.columns.get_loc('missing mortality_race2_black') + 1)
    write_into_mortality_race2_black.value = entire_mortality_race2_black

    write_into_mortality_race2_other = ws.cell(row=row_index,
                                               column=pandas_type.columns.get_loc('missing mortality_race2_other') + 1)
    write_into_mortality_race2_other.value = entire_mortality_race2_other

    write_into_mortality_race2_unknown = ws.cell(row=row_index, column=pandas_type.columns.get_loc(
        'missing mortality_race2_unknown') + 1)
    write_into_mortality_race2_unknown.value = entire_mortality_race2_unknown



for individual in states:
    filepath = os.path.join('QC Folder', individual)
    os.chdir(filepath)
    if check_authenticity(individual, 'State-Year'):
        os.chdir("/Users/kennny/Desktop/Kenny's QC 2025")
        print(f'{individual} has error')
        continue
    state_year_general(individual)

    os.chdir("/Users/kennny/Desktop/Kenny's QC 2025")
wb.save('QC Standard File.xlsx')